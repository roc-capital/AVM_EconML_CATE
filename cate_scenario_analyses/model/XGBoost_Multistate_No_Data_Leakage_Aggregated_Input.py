"""
COMPACT STRATIFIED AVM MODEL - Dynamic State Handling
8 PRICE TIERS | QUANTILE REGRESSION | MAXIMUM PERFORMANCE
"""

import pandas as pd
import numpy as np
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.cluster import MiniBatchKMeans
import warnings;

warnings.filterwarnings('ignore')
import time, os

os.environ['OMP_NUM_THREADS'] = '1'

# -------------------------
# CONFIG
# -------------------------
Y_COL, PROPERTYID_COL, STATE_COL = 'sale_price', 'property_id', 'state'
MIN_PRICE_THRESHOLD, TEST_SIZE, RANDOM_STATE, N_JOBS, N_GEO_CLUSTERS = 100000, 0.2, 42, -1, 8
PARALLEL_QUANTILES, USE_MEMORY_OPTIMIZATION, REDUCED_ESTIMATORS = True, True, True

UNIFIED_DATA_PATH = "/Users/jenny.lin/BASIS_AVM_Onboarding/cate_scenario_analyses/data/state_specific_parquet_files_v2/Aggregated_States_2025.parquet"
OUTPUT_DIR = "/Users/jenny.lin/BASIS_AVM_Onboarding/cate_scenario_analyses/model_outputs"

QUANTILES = [0.1, 0.5, 0.9]
PRICE_TIERS = {
    'very_low': (0, 200000), 'low': (200000, 300000), 'lower_mid': (300000, 400000),
    'mid': (400000, 500000), 'upper_mid': (500000, 650000), 'high': (650000, 850000),
    'very_high': (850000, 1200000), 'ultra_high': (1200000, np.inf)
}

N_ESTIMATORS, EARLY_STOPPING = (500, 50) if REDUCED_ESTIMATORS else (800, 75)

BASE_FEATURES = ["living_sqft", "lot_sqft", "year_built", "bedrooms", "full_baths", "garage_spaces", "geo_cluster"]
INCOME_FEATURES = ["median_earnings_total", "pct_white"]
POLITICAL_FEATURES = ["per_point_diff"]
NEIGHBORHOOD_FEATURES = [
    "nbhd_population", "nbhd_pop_density", "nbhd_median_age", "nbhd_household_size",
    "nbhd_pct_age_0_5", "nbhd_pct_age_6_11", "nbhd_pct_age_12_17", "nbhd_pct_children",
    "nbhd_median_income", "nbhd_avg_income", "nbhd_per_capita_income", "nbhd_poverty_rate",
    "nbhd_pct_high_income", "nbhd_cost_of_living_index", "nbhd_housing_cost_index",
    "nbhd_median_home_value", "nbhd_median_rent", "nbhd_homeownership_rate",
    "nbhd_median_year_built", "nbhd_vacancy_rate", "nbhd_pct_bachelors",
    "nbhd_pct_grad_degree", "nbhd_pct_high_school", "nbhd_crime_index",
    "nbhd_air_quality_index", "nbhd_ozone_index", "nbhd_median_commute_min",
    "nbhd_pct_work_from_home", "nbhd_pct_drive_alone", "nbhd_pct_public_transit",
    "nbhd_pct_white_collar", "nbhd_pct_professional", "nbhd_pct_healthcare",
    "nbhd_pct_new_housing", "nbhd_pct_single_family", "nbhd_pct_large_apartments",
    "nbhd_pct_married_families", "nbhd_pct_families_with_children"
]
SCHOOL_FEATURES = [
    "nbhd_avg_student_teacher_ratio", "nbhd_avg_school_size", "nbhd_above_avg_schools_cnt",
    "nbhd_elementary_schools_cnt", "nbhd_middle_schools_cnt", "nbhd_high_schools_cnt",
    "nbhd_avg_pct_free_reduced_lunch", "nbhd_ap_schools_cnt", "nbhd_gifted_prog_schools_cnt",
    "nbhd_avg_college_going_rate"
]
DISTRICT_FEATURES = ["district_per_pupil_spending", "district_pct_spending_instruction",
                     "district_teachers_per_1000_students"]
ENGINEERED_FEATURES = ["sqft_per_bedroom", "lot_to_living_ratio", "property_age", "is_new", "has_garage",
                       "luxury_score", "log_sqft", "age_squared", "income_education_score", "affordability_ratio"]
CLUSTER_FEATURES = ["cluster_avg_price", "cluster_med_price"]


# -------------------------
# HELPER FUNCTIONS
# -------------------------

def get_all_defined_features():
    """Return comprehensive list of all feature columns defined in the script."""
    all_features = (
            BASE_FEATURES +
            INCOME_FEATURES +
            POLITICAL_FEATURES +
            NEIGHBORHOOD_FEATURES +
            SCHOOL_FEATURES +
            DISTRICT_FEATURES +
            ENGINEERED_FEATURES +
            CLUSTER_FEATURES
    )
    return sorted(list(set(all_features)))  # Remove duplicates and sort


def print_all_features():
    """Print all defined features organized by category."""
    print("\n" + "=" * 60)
    print("ALL DEFINED FEATURES BY CATEGORY")
    print("=" * 60)
    print(f"\nBASE FEATURES ({len(BASE_FEATURES)}):")
    for f in BASE_FEATURES: print(f"  - {f}")
    print(f"\nINCOME FEATURES ({len(INCOME_FEATURES)}):")
    for f in INCOME_FEATURES: print(f"  - {f}")
    print(f"\nPOLITICAL FEATURES ({len(POLITICAL_FEATURES)}):")
    for f in POLITICAL_FEATURES: print(f"  - {f}")
    print(f"\nNEIGHBORHOOD FEATURES ({len(NEIGHBORHOOD_FEATURES)}):")
    for f in NEIGHBORHOOD_FEATURES: print(f"  - {f}")
    print(f"\nSCHOOL FEATURES ({len(SCHOOL_FEATURES)}):")
    for f in SCHOOL_FEATURES: print(f"  - {f}")
    print(f"\nDISTRICT FEATURES ({len(DISTRICT_FEATURES)}):")
    for f in DISTRICT_FEATURES: print(f"  - {f}")
    print(f"\nENGINEERED FEATURES ({len(ENGINEERED_FEATURES)}):")
    for f in ENGINEERED_FEATURES: print(f"  - {f}")
    print(f"\nCLUSTER FEATURES ({len(CLUSTER_FEATURES)}):")
    for f in CLUSTER_FEATURES: print(f"  - {f}")
    print(f"\nTOTAL UNIQUE FEATURES: {len(get_all_defined_features())}")
    print("=" * 60 + "\n")


def optimize_dtypes(df):
    """Reduce memory usage."""
    for col in df.select_dtypes(include=['float64']).columns: df[col] = df[col].astype('float32')
    for col in df.select_dtypes(include=['int64']).columns: df[col] = df[col].astype('int32')
    return df


def feature_importance(models, feature_names, metrics):
    """Calculate weighted feature importance."""
    rows = []
    for tier, model_dict in models.items():
        booster = model_dict['q50'].get_booster()
        scores = booster.get_score(importance_type="gain")
        weight = metrics[tier]["n_test"]
        for k, v in scores.items():
            idx = int(k[1:])
            if idx < len(feature_names): rows.append((feature_names[idx], v, weight))

    if not rows: return pd.DataFrame(columns=["feature", "importance"])

    df = pd.DataFrame(rows, columns=["feature", "gain", "weight"])
    out = df.assign(weighted_gain=df["gain"] * df["weight"]).groupby("feature", as_index=False).agg(
        total_gain=("weighted_gain", "sum")).sort_values("total_gain", ascending=False)
    out["importance"] = out["total_gain"] / out["total_gain"].sum()
    return out[["feature", "importance"]].head(50)


def load_unified_data(filepath):
    """Load data and discover states dynamically."""
    print(f"Loading: {filepath}")
    df = pd.read_parquet(filepath, engine='pyarrow')
    df.columns = df.columns.str.lower()
    print(f"Records: {len(df):,} | Memory: {df.memory_usage(deep=True).sum() / 1024 ** 2:.1f} MB")

    if STATE_COL not in df.columns:
        raise ValueError(f"'{STATE_COL}' column not found!")

    state_counts = df[STATE_COL].value_counts().sort_values(ascending=False)
    print(f"\n{'=' * 60}\nSTATES DISCOVERED: {len(state_counts)}")
    for state, count in state_counts.items(): print(f"  {state}: {count:,}")

    MIN_RECORDS = 100
    states = [s for s, c in state_counts.items() if c >= MIN_RECORDS]
    if len(states) < len(state_counts):
        print(f"⚠️  Excluding states with <{MIN_RECORDS} records")

    print(f"✓ Processing {len(states)} states: {states}\n{'=' * 60}\n")
    return optimize_dtypes(df[df[STATE_COL].isin(states)]), states


def discover_features(df, feature_groups):
    """Find available features."""
    all_features = [f for group in feature_groups for f in group]
    available = [f for f in all_features if f in df.columns]
    missing = len(all_features) - len(available)
    print(f"Features: {len(available)}/{len(all_features)} available" + (f" ({missing} missing)" if missing else ""))
    return available


def engineer_features(df):
    """Create engineered features."""
    if 'living_sqft' in df.columns and 'bedrooms' in df.columns:
        df['sqft_per_bedroom'] = df['living_sqft'] / (df['bedrooms'] + 1)
    if 'lot_sqft' in df.columns and 'living_sqft' in df.columns:
        df['lot_to_living_ratio'] = df['lot_sqft'] / (df['living_sqft'] + 1)
    if 'year_built' in df.columns:
        df['property_age'] = 2024 - df['year_built']
        df['is_new'] = (df['property_age'] <= 5).astype(int)
        df['age_squared'] = df['property_age'] ** 2
    if 'garage_spaces' in df.columns:
        df['has_garage'] = (df['garage_spaces'] > 0).astype(int)
    if 'living_sqft' in df.columns:
        df['log_sqft'] = np.log1p(df['living_sqft'])

    luxury = []
    if 'living_sqft' in df.columns: luxury.append(df['living_sqft'] / 1000)
    if 'full_baths' in df.columns: luxury.append(df['full_baths'])
    if 'garage_spaces' in df.columns: luxury.append(df['garage_spaces'])
    if luxury: df['luxury_score'] = sum(luxury) / len(luxury)

    if 'nbhd_median_income' in df.columns and 'nbhd_pct_bachelors' in df.columns:
        df['income_education_score'] = df['nbhd_median_income'] * df['nbhd_pct_bachelors']
    if 'nbhd_median_income' in df.columns and 'nbhd_median_home_value' in df.columns:
        df['affordability_ratio'] = df['nbhd_median_income'] / (df['nbhd_median_home_value'] + 1)
    return df


def create_geo_clusters(df):
    """Create geographic clusters."""
    if not all(c in df.columns for c in ['latitude', 'longitude']):
        df['geo_cluster'] = 0
        return df

    valid = df[['latitude', 'longitude']].notna().all(axis=1)
    if valid.sum() < N_GEO_CLUSTERS:
        df['geo_cluster'] = 0
        return df

    df['geo_cluster'] = 0
    kmeans = MiniBatchKMeans(n_clusters=N_GEO_CLUSTERS, random_state=RANDOM_STATE, batch_size=1000, n_init=3)
    df.loc[valid, 'geo_cluster'] = kmeans.fit_predict(df.loc[valid, ['latitude', 'longitude']])
    return df


def add_cluster_features(df):
    """Add cluster aggregates."""
    if 'geo_cluster' not in df.columns or Y_COL not in df.columns:
        df['cluster_avg_price'] = df[Y_COL].median() if Y_COL in df.columns else 0
        df['cluster_med_price'] = df[Y_COL].median() if Y_COL in df.columns else 0
        return df

    stats = df.groupby('geo_cluster')[Y_COL].agg(['mean', 'median']).reset_index()
    stats.columns = ['geo_cluster', 'cluster_avg_price', 'cluster_med_price']
    df = df.merge(stats, on='geo_cluster', how='left')
    df[['cluster_avg_price', 'cluster_med_price']] = df[['cluster_avg_price', 'cluster_med_price']].fillna(
        df[Y_COL].median())
    return df


def prepare_state_data(df, state_name):
    """Prepare data for modeling."""
    print(f"\n{'=' * 60}\n{state_name}")
    state_df = df[df[STATE_COL] == state_name].copy()
    state_df = state_df[state_df[Y_COL] >= MIN_PRICE_THRESHOLD]
    print(f"Records: {len(state_df):,}")

    if len(state_df) < 100: return None, None

    state_df = engineer_features(create_geo_clusters(state_df))
    state_df = add_cluster_features(state_df)
    state_df['price_tier'] = state_df[Y_COL].apply(
        lambda p: next((t for t, (l, h) in PRICE_TIERS.items() if l <= p < h), 'ultra_high'))

    features = discover_features(state_df, [BASE_FEATURES, INCOME_FEATURES, POLITICAL_FEATURES,
                                            NEIGHBORHOOD_FEATURES, SCHOOL_FEATURES, DISTRICT_FEATURES,
                                            ENGINEERED_FEATURES, CLUSTER_FEATURES])

    cols = features + [Y_COL, PROPERTYID_COL, 'price_tier']
    state_df = state_df[list(dict.fromkeys(cols))].copy()
    state_df[features] = state_df[features].fillna(state_df[features].median())
    state_df = state_df.dropna(subset=[Y_COL])

    print(f"Final: {len(state_df):,} records, {len(features)} features")
    return state_df, features


def train_quantile_model(X_train, y_train, X_test, y_test, quantile):
    """Train single quantile model."""
    model = XGBRegressor(objective='reg:quantileerror', quantile_alpha=quantile, n_estimators=N_ESTIMATORS,
                         learning_rate=0.05, max_depth=6, min_child_weight=3, subsample=0.8,
                         colsample_bytree=0.8, random_state=RANDOM_STATE, n_jobs=N_JOBS, tree_method='hist')
    model.fit(X_train, y_train, verbose=False)
    return model


def process_single_state(state_name, df):
    """Process one state."""
    start = time.time()
    state_df, features = prepare_state_data(df, state_name)
    if state_df is None: return None

    models, metrics, predictions_list = {}, {}, []

    for tier_name, (low, high) in PRICE_TIERS.items():
        tier_df = state_df[state_df['price_tier'] == tier_name]
        if len(tier_df) < 50: continue

        print(f"\n  {tier_name} (${low:,}-${high:,}): {len(tier_df):,} samples")

        X, y, ids = tier_df[features].values, tier_df[Y_COL].values, tier_df[PROPERTYID_COL].values
        X_train, X_test, y_train, y_test, ids_train, ids_test = train_test_split(
            X, y, ids, test_size=min(TEST_SIZE, 0.3) if len(tier_df) < 200 else TEST_SIZE, random_state=RANDOM_STATE)

        tier_models, tier_preds = {}, []
        for q in QUANTILES:
            q_label = f"q{int(q * 100)}"
            model = train_quantile_model(X_train, y_train, X_test, y_test, q)
            tier_models[q_label] = model
            tier_preds.append(model.predict(X_test))

        models[tier_name] = tier_models
        y_pred = tier_preds[1]  # median
        mae, mape, r2 = mean_absolute_error(y_test, y_pred), np.mean(
            np.abs((y_test - y_pred) / y_test)) * 100, r2_score(y_test, y_pred)
        coverage = np.mean((y_test >= tier_preds[0]) & (y_test <= tier_preds[2])) * 100

        metrics[tier_name] = {'n_train': len(X_train), 'n_test': len(X_test), 'mae': mae, 'mape': mape, 'r2': r2,
                              'coverage_80': coverage}
        print(f"    MAE: ${mae:,.0f} | MAPE: {mape:.2f}% | R²: {r2:.4f} | Coverage: {coverage:.1f}%")

        predictions_list.append(pd.DataFrame({
            'property_id': ids_test, 'actual': y_test, 'predicted': y_pred,
            'pred_lower': tier_preds[0], 'pred_upper': tier_preds[2],
            'price_tier': tier_name, 'state': state_name
        }))

    if not models: return None

    return {
        'state': state_name, 'models': models, 'metrics': metrics,
        'predictions': pd.concat(predictions_list, ignore_index=True),
        'feature_importance': feature_importance(models, features, metrics),
        'feature_names': features, 'time': time.time() - start
    }


def main():
    """Main execution."""
    print(f"\n{'=' * 60}\nDYNAMIC STRATIFIED AVM\n{'=' * 60}")

    df, states = load_unified_data(UNIFIED_DATA_PATH)
    all_results = [r for state in states if (r := process_single_state(state, df))]

    print(f"\n{'=' * 60}\nSUMMARY: {len(all_results)}/{len(states)} states processed")
    for r in all_results:
        print(f"\n{r['state']} ({r['time']:.1f}s):")
        for tier, m in r['metrics'].items():
            print(f"  {tier}: MAE=${m['mae']:,.0f}, R²={m['r2']:.3f}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Save individual state outputs
    for r in all_results:
        state_clean = r['state'].replace(' ', '_').lower()
        r['predictions'].to_csv(f"{OUTPUT_DIR}/{state_clean}_predictions.csv", index=False)
        r['feature_importance'].to_csv(f"{OUTPUT_DIR}/{state_clean}_feature_importance.csv", index=False)
        pd.DataFrame(r['metrics']).T.to_csv(f"{OUTPUT_DIR}/{state_clean}_metrics.csv")

    # Generate comprehensive Excel report
    if all_results:
        generate_excel_report(all_results)

    print(f"\n✓ Complete! Outputs in {OUTPUT_DIR}")


def generate_excel_report(all_results):
    """Generate comprehensive Excel workbook with multiple analysis tabs."""
    print(f"\n{'=' * 60}\nCREATING COMPREHENSIVE EXCEL REPORT\n{'=' * 60}")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    # Combine all predictions
    all_preds = pd.concat([r['predictions'] for r in all_results], ignore_index=True)

    # TAB 1: Executive Summary
    ws = wb.create_sheet("Executive Summary", 0)
    ws['A1'] = 'STRATIFIED AVM MODEL - QUANTILE REGRESSION'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    ws['A2'] = f'Generated: {time.strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:H2')

    overall_r2 = r2_score(all_preds['actual'], all_preds['predicted'])
    overall_mae = mean_absolute_error(all_preds['actual'], all_preds['predicted'])
    overall_mape = np.mean(np.abs((all_preds['actual'] - all_preds['predicted']) / all_preds['actual']) * 100)

    ws['A4'] = 'OVERALL PERFORMANCE'
    ws['A4'].font = Font(bold=True, size=12)
    summary_data = [
        ['Metric', 'Value'],
        ['Total Properties', len(all_preds)],
        ['Overall R²', overall_r2],
        ['Overall MAE', f'${overall_mae:,.0f}'],
        ['Overall MAPE (%)', f'{overall_mape:.2f}%'],
        ['States Processed', len(all_results)],
        ['Price Tiers', len(PRICE_TIERS)],
    ]
    for row_idx, (label, value) in enumerate(summary_data, start=5):
        ws[f'A{row_idx}'] = label
        ws[f'A{row_idx}'].font = Font(bold=True)
        ws[f'B{row_idx}'] = value
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # TAB 2: State Comparison
    ws_states = wb.create_sheet("State Comparison")
    state_rows = []
    for r in all_results:
        preds = r['predictions']
        state_rows.append({
            'State': r['state'],
            'N Properties': len(preds),
            'R²': r2_score(preds['actual'], preds['predicted']),
            'MAE': mean_absolute_error(preds['actual'], preds['predicted']),
            'MAPE (%)': np.mean(np.abs((preds['actual'] - preds['predicted']) / preds['actual']) * 100),
            'Min Price': preds['actual'].min(),
            'Max Price': preds['actual'].max(),
            'Mean Price': preds['actual'].mean(),
            'Median Price': preds['actual'].median(),
        })

    state_df = pd.DataFrame(state_rows)
    headers = list(state_df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_states.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for row_idx, row_data in enumerate(state_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_states.cell(row=row_idx, column=col_idx, value=value)

    for col in ws_states.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_states.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)

    # TAB 3: Tier Comparison
    ws_tiers = wb.create_sheet("Tier Comparison")
    tier_rows = []
    for r in all_results:
        for tier, m in r['metrics'].items():
            tier_rows.append({
                'State': r['state'],
                'Tier': tier,
                'N Train': m['n_train'],
                'N Test': m['n_test'],
                'R²': m['r2'],
                'MAE': m['mae'],
                'MAPE (%)': m['mape'],
                'Coverage 80%': m['coverage_80'],
            })

    tier_df = pd.DataFrame(tier_rows)
    for col_idx, header in enumerate(tier_df.columns, start=1):
        cell = ws_tiers.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for row_idx, row_data in enumerate(tier_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_tiers.cell(row=row_idx, column=col_idx, value=value)

    for col in ws_tiers.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_tiers.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

    # TAB 4: Feature Importance (Top 30 aggregated)
    ws_feat = wb.create_sheet("Feature Importance")
    all_fi = pd.concat([r['feature_importance'] for r in all_results])
    avg_fi = all_fi.groupby('feature')['importance'].mean().reset_index()
    avg_fi = avg_fi.sort_values('importance', ascending=False).head(30)
    avg_fi.insert(0, 'rank', range(1, len(avg_fi) + 1))

    for col_idx, header in enumerate(['Rank', 'Feature', 'Importance'], start=1):
        cell = ws_feat.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for row_idx, row_data in enumerate(avg_fi.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_feat.cell(row=row_idx, column=col_idx, value=value)

    ws_feat.column_dimensions['A'].width = 8
    ws_feat.column_dimensions['B'].width = 35
    ws_feat.column_dimensions['C'].width = 15

    # TAB 5: Sample Predictions (1000 per state, max 10k total)
    ws_sample = wb.create_sheet("Sample Predictions")
    samples = all_preds.groupby('state').head(1000).head(10000)
    samples = samples[['state', 'property_id', 'price_tier', 'actual', 'predicted',
                       'pred_lower', 'pred_upper']]

    for col_idx, header in enumerate(samples.columns, start=1):
        cell = ws_sample.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for row_idx, row_data in enumerate(samples.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_sample.cell(row=row_idx, column=col_idx, value=value)

    for col in ws_sample.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws_sample.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

    # Save workbook
    excel_path = f"{OUTPUT_DIR}/comprehensive_model_analysis.xlsx"
    wb.save(excel_path)

    print(f"✓ Excel report: {excel_path}")
    print(f"  Tabs: Executive Summary, State Comparison, Tier Comparison,")
    print(f"        Feature Importance, Sample Predictions")

    # Also save summary CSVs
    state_df.to_csv(f"{OUTPUT_DIR}/state_comparison.csv", index=False)
    tier_df.to_csv(f"{OUTPUT_DIR}/tier_comparison.csv", index=False)
    print(f"✓ Summary CSVs saved")
    print('=' * 60)


if __name__ == "__main__": main()
